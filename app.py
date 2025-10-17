from flask import Flask, render_template, request, jsonify, send_file
from flask_cors import CORS
from datetime import datetime
import sqlite3
import csv
import io
import qrcode
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

app = Flask(__name__)
CORS(app)

DATABASE = 'quilometragem.db'

def get_db():
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    db = get_db()
    db.execute(
        """
        CREATE TABLE IF NOT EXISTS registros (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            condutor TEXT NOT NULL,
            placa_veiculo TEXT NOT NULL,
            data_saida TEXT NOT NULL,
            data_chegada TEXT NOT NULL,
            km_inicial REAL NOT NULL,
            km_final REAL NOT NULL,
            distancia_percorrida REAL NOT NULL,
            observacoes TEXT,
            data_criacao TEXT NOT NULL,
            data_atualizacao TEXT
        )
        """
    )
    db.execute(
        """
        CREATE TABLE IF NOT EXISTS historico_alteracoes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            registro_id INTEGER NOT NULL,
            campo_alterado TEXT NOT NULL,
            valor_anterior TEXT,
            valor_novo TEXT,
            usuario_alteracao TEXT,
            data_alteracao TEXT NOT NULL,
            FOREIGN KEY (registro_id) REFERENCES registros (id)
        )
        """
    )
    db.commit()

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/registros', methods=['GET'])
def listar_registros():
    try:
        db = get_db()
        condutor = request.args.get('condutor', '')
        placa = request.args.get('placa', '')
        data_inicio = request.args.get('data_inicio', '')
        data_fim = request.args.get('data_fim', '')

        query = 'SELECT * FROM registros WHERE 1=1'
        params = []
        if condutor:
            query += ' AND LOWER(condutor) LIKE LOWER(?)'
            params.append(f'%{condutor}%')
        if placa:
            query += ' AND LOWER(placa_veiculo) LIKE LOWER(?)'
            params.append(f'%{placa}%')
        if data_inicio:
            query += ' AND data_saida >= ?'
            params.append(data_inicio)
        if data_fim:
            query += ' AND data_chegada <= ?'
            params.append(data_fim)
        query += ' ORDER BY data_saida DESC'
        registros = db.execute(query, params).fetchall()
        return jsonify([dict(r) for r in registros])
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/registros', methods=['POST'])
def criar_registro():
    try:
        dados = request.json
        obrig = ['condutor', 'placa_veiculo', 'data_saida', 'data_chegada', 'km_inicial', 'km_final']
        for c in obrig:
            if not dados.get(c):
                return jsonify({'error': f'Campo obrigatório: {c}'}), 400
        km_i = float(dados['km_inicial'])
        km_f = float(dados['km_final'])
        if km_f < km_i:
            return jsonify({'error': 'KM final deve ser maior que KM inicial'}), 400
        dist = km_f - km_i
        db = get_db()
        cur = db.execute(
            'INSERT INTO registros (condutor, placa_veiculo, data_saida, data_chegada, km_inicial, km_final, distancia_percorrida, observacoes, data_criacao) VALUES (?,?,?,?,?,?,?,?,?)',
            (
                dados['condutor'], dados['placa_veiculo'], dados['data_saida'], dados['data_chegada'],
                km_i, km_f, dist, dados.get('observacoes', ''), datetime.now().isoformat()
            )
        )
        db.commit()
        return jsonify({'id': cur.lastrowid, 'message': 'Registro criado com sucesso', 'distancia_percorrida': dist}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/registros/<int:id>', methods=['GET'])
def obter_registro(id):
    db = get_db()
    r = db.execute('SELECT * FROM registros WHERE id=?', (id,)).fetchone()
    if not r:
        return jsonify({'error': 'Registro não encontrado'}), 404
    return jsonify(dict(r))

@app.route('/api/registros/<int:id>', methods=['PUT'])
def atualizar_registro(id):
    try:
        dados = request.json
        db = get_db()
        atual = db.execute('SELECT * FROM registros WHERE id=?', (id,)).fetchone()
        if not atual:
            return jsonify({'error': 'Registro não encontrado'}), 404
        km_i = float(dados.get('km_inicial', atual['km_inicial']))
        km_f = float(dados.get('km_final', atual['km_final']))
        if km_f < km_i:
            return jsonify({'error': 'KM final deve ser maior que KM inicial'}), 400
        dist = km_f - km_i
        campos = {'condutor':'Condutor','placa_veiculo':'Placa do Veículo','data_saida':'Data de Saída','data_chegada':'Data de Chegada','km_inicial':'KM Inicial','km_final':'KM Final','observacoes':'Observações'}
        for campo, nome in campos.items():
            novo = dados.get(campo)
            if novo is not None and str(novo) != str(atual[campo]):
                db.execute(
                    'INSERT INTO historico_alteracoes (registro_id, campo_alterado, valor_anterior, valor_novo, usuario_alteracao, data_alteracao) VALUES (?,?,?,?,?,?)',
                    (id, nome, str(atual[campo]), str(novo), dados.get('usuario_alteracao', 'Sistema'), datetime.now().isoformat())
                )
        db.execute(
            'UPDATE registros SET condutor=?, placa_veiculo=?, data_saida=?, data_chegada=?, km_inicial=?, km_final=?, distancia_percorrida=?, observacoes=?, data_atualizacao=? WHERE id=?',
            (
                dados.get('condutor', atual['condutor']),
                dados.get('placa_veiculo', atual['placa_veiculo']),
                dados.get('data_saida', atual['data_saida']),
                dados.get('data_chegada', atual['data_chegada']),
                km_i, km_f, dist,
                dados.get('observacoes', atual['observacoes']),
                datetime.now().isoformat(), id
            )
        )
        db.commit()
        return jsonify({'message':'Registro atualizado com sucesso','distancia_percorrida':dist})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/registros/<int:id>', methods=['DELETE'])
def deletar_registro(id):
    db = get_db()
    db.execute('DELETE FROM registros WHERE id=?', (id,))
    db.execute('DELETE FROM historico_alteracoes WHERE registro_id=?', (id,))
    db.commit()
    return jsonify({'message':'Registro deletado com sucesso'})

@app.route('/api/registros/<int:id>/historico', methods=['GET'])
def obter_historico(id):
    db = get_db()
    h = db.execute('SELECT * FROM historico_alteracoes WHERE registro_id=? ORDER BY data_alteracao DESC', (id,)).fetchall()
    return jsonify([dict(x) for x in h])

@app.route('/api/exportar/csv', methods=['GET'])
def exportar_csv():
    db = get_db()
    regs = db.execute('SELECT * FROM registros ORDER BY data_saida DESC').fetchall()
    out = io.StringIO()
    w = csv.writer(out)
    w.writerow(['ID','Condutor','Placa','Data Saída','Data Chegada','KM Inicial','KM Final','Distância (KM)','Observações'])
    for r in regs:
        w.writerow([r['id'], r['condutor'], r['placa_veiculo'], r['data_saida'], r['data_chegada'], r['km_inicial'], r['km_final'], r['distancia_percorrida'], r['observacoes']])
    out.seek(0)
    return send_file(io.BytesIO(out.getvalue().encode('utf-8-sig')), mimetype='text/csv', as_attachment=True, download_name='quilometragem.csv')

@app.route('/api/exportar/excel', methods=['GET'])
def exportar_excel():
    db = get_db()
    regs = db.execute('SELECT * FROM registros ORDER BY data_saida DESC').fetchall()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Quilometragem'
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    headers = ['ID','Condutor','Placa','Data Saída','Data Chegada','KM Inicial','KM Final','Distância (KM)','Observações']
    for i,h in enumerate(headers,1):
        c = ws.cell(row=1, column=i, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal='center', vertical='center')
    for row, r in enumerate(regs, 2):
        ws.cell(row=row, column=1, value=r['id'])
        ws.cell(row=row, column=2, value=r['condutor'])
        ws.cell(row=row, column=3, value=r['placa_veiculo'])
        ws.cell(row=row, column=4, value=r['data_saida'])
        ws.cell(row=row, column=5, value=r['data_chegada'])
        ws.cell(row=row, column=6, value=r['km_inicial'])
        ws.cell(row=row, column=7, value=r['km_final'])
        ws.cell(row=row, column=8, value=r['distancia_percorrida'])
        ws.cell(row=row, column=9, value=r['observacoes'])
    for col in ws.columns:
        maxlen = 0
        col_letter = col[0].column_letter
        for cell in col:
            maxlen = max(maxlen, len(str(cell.value)) if cell.value is not None else 0)
        ws.column_dimensions[col_letter].width = min(maxlen+2, 50)
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(bio, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name='quilometragem.xlsx')

@app.route('/api/exportar/pdf', methods=['GET'])
def exportar_pdf():
    db = get_db()
    regs = db.execute('SELECT * FROM registros ORDER BY data_saida DESC').fetchall()
    bio = io.BytesIO()
    doc = SimpleDocTemplate(bio, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    elements.append(Paragraph('Relatório de Quilometragem', styles['Heading1']))
    elements.append(Spacer(1, 0.3*inch))
    data = [['ID','Condutor','Placa','Saída','Chegada','KM Ini','KM Fim','Dist.']]
    for r in regs:
        data.append([str(r['id']), r['condutor'][:15], r['placa_veiculo'], r['data_saida'], r['data_chegada'], str(r['km_inicial']), str(r['km_final']), str(r['distancia_percorrida'])])
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.grey),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 10),
        ('BOTTOMPADDING', (0,0), (-1,0), 12),
        ('BACKGROUND', (0,1), (-1,-1), colors.beige),
        ('GRID', (0,0), (-1,-1), 1, colors.black),
        ('FONTSIZE', (0,1), (-1,-1), 8),
    ]))
    elements.append(table)
    doc.build(elements)
    bio.seek(0)
    return send_file(bio, mimetype='application/pdf', as_attachment=True, download_name='quilometragem.pdf')

@app.route('/api/qrcode', methods=['GET'])
def gerar_qrcode():
    url = request.args.get('url', request.host_url)
    qr = qrcode.QRCode(version=1, error_correction=qrcode.constants.ERROR_CORRECT_L, box_size=10, border=4)
    qr.add_data(url)
    qr.make(fit=True)
    img = qr.make_image(fill_color='black', back_color='white')
    bio = io.BytesIO()
    img.save(bio, format='PNG')
    bio.seek(0)
    return send_file(bio, mimetype='image/png')

if __name__ == '__main__':
    init_db()
    app.run(debug=True, host='0.0.0.0', port=5000)
